import pandas as pd
import os
import logging
from typing import Dict, Optional
from openpyxl import load_workbook

class ExcelService:
    def __init__(self, config):
        self.config = config
        self.logger = logging.getLogger(self.__class__.__name__)
        self.excel_path = config.EXCEL_USER_DATABASE
        self.columns = config.EXCEL_COLUMNS

        # Validate Excel file exists
        if not os.path.exists(self.excel_path):
            self.logger.error(f"Excel database file not found: {self.excel_path}")
            raise FileNotFoundError(f"Excel database file not found: {self.excel_path}")

    def get_user_info(self, card_uid: str) -> Optional[Dict[str, str]]:
        try:
            self.logger.info(f"Looking up user info for card UID: {card_uid}")

            # Check all configured worksheets
            for worksheet_name in self.config.EXCEL_WORKSHEETS:
                user_info = self._search_worksheet(worksheet_name, card_uid)
                if user_info:
                    return user_info

            self.logger.info(f"No user found for card UID: {card_uid}")
            return None

        except Exception as e:
            self.logger.error(f"Error reading Excel database: {str(e)}")
            return None

    def _search_worksheet(self, worksheet_name: str, card_uid: str) -> Optional[Dict[str, str]]:
        try:
            self.logger.debug(f"Searching worksheet '{worksheet_name}' for UID: {card_uid}")

            # Read specific worksheet
            df = pd.read_excel(self.excel_path, sheet_name=worksheet_name, header=None)

            self.logger.debug(f"Worksheet '{worksheet_name}' shape: {df.shape}")

            # Convert column letters to indices (A=0, B=1, D=3, E=4, K=10)
            col_indices = {
                'supervisor': self._column_letter_to_index(self.columns['supervisor']),
                'gender': self._column_letter_to_index(self.columns['gender']),
                'firstname': self._column_letter_to_index(self.columns['firstname']),
                'lastname': self._column_letter_to_index(self.columns['lastname']),
                'card_uid': self._column_letter_to_index(self.columns['card_uid'])
            }

            self.logger.debug(f"Column indices: {col_indices}")

            # Ensure we have enough columns
            max_col = max(col_indices.values())
            if df.shape[1] <= max_col:
                self.logger.warning(f"Worksheet '{worksheet_name}' doesn't have enough columns (has {df.shape[1]}, need {max_col + 1})")
                return None

            # Convert UID column to string and strip whitespace
            uid_col_idx = col_indices['card_uid']
            df.iloc[:, uid_col_idx] = df.iloc[:, uid_col_idx].astype(str).str.strip()

            # Search for matching UID (case-insensitive)
            mask = df.iloc[:, uid_col_idx].str.lower() == card_uid.lower()
            matching_rows = df[mask]

            if matching_rows.empty:
                return None

            if len(matching_rows) > 1:
                self.logger.warning(f"Multiple users found for card UID: {card_uid} in worksheet '{worksheet_name}'. Using first match.")

            # Get first matching row
            user_row = matching_rows.iloc[0]

            # Extract user information
            supervisor = str(user_row.iloc[col_indices['supervisor']]).strip() if pd.notna(user_row.iloc[col_indices['supervisor']]) else ''
            gender = str(user_row.iloc[col_indices['gender']]).strip() if pd.notna(user_row.iloc[col_indices['gender']]) else ''
            firstname = str(user_row.iloc[col_indices['firstname']]).strip() if pd.notna(user_row.iloc[col_indices['firstname']]) else ''
            lastname = str(user_row.iloc[col_indices['lastname']]).strip() if pd.notna(user_row.iloc[col_indices['lastname']]) else ''

            # Check for guest card
            is_guest_card = lastname.lower() == "gästekarte"

            # Combine names
            if firstname and lastname:
                full_name = f"{firstname} {lastname}"
            elif firstname:
                full_name = firstname
            elif lastname:
                full_name = lastname
            else:
                full_name = "Unbekannt"

            user_info = {
                'card_uid': str(user_row.iloc[col_indices['card_uid']]).strip(),
                'name': full_name,
                'firstname': firstname,
                'lastname': lastname,
                'supervisor': supervisor,
                'gender': gender,
                'email': self._generate_outlook_address(lastname, firstname) if not is_guest_card else '',
                'supervisor_email': supervisor,  # Already in "Nachname, Vorname" format
                'is_guest_card': is_guest_card
            }

            if is_guest_card:
                self.logger.info(f"Found guest card: {card_uid} - Supervisor: {supervisor}")
            else:
                self.logger.info(f"Found user: {full_name} ({supervisor}) in worksheet '{worksheet_name}'")

            return user_info

        except Exception as e:
            self.logger.error(f"Error searching worksheet '{worksheet_name}': {str(e)}")
            return None

    def delete_user(self, card_uid: str) -> bool:
        try:
            self.logger.info(f"Deleting user from Excel database: {card_uid}")

            # Create backup first
            backup_path = f"{self.excel_path}.backup"
            import shutil
            shutil.copy2(self.excel_path, backup_path)
            self.logger.info(f"Created backup: {backup_path}")

            # Load workbook to preserve formatting
            workbook = load_workbook(self.excel_path)

            total_deleted = 0

            # Check all configured worksheets
            for worksheet_name in self.config.EXCEL_WORKSHEETS:
                try:
                    if worksheet_name in workbook.sheetnames:
                        worksheet = workbook[worksheet_name]
                        deleted_count = self._delete_from_worksheet(worksheet, card_uid)
                        total_deleted += deleted_count
                        self.logger.info(f"Deleted {deleted_count} row(s) from worksheet '{worksheet_name}'")
                    else:
                        self.logger.warning(f"Worksheet '{worksheet_name}' not found in workbook")
                except Exception as e:
                    self.logger.error(f"Error deleting from worksheet '{worksheet_name}': {str(e)}")

            if total_deleted > 0:
                # Save workbook
                workbook.save(self.excel_path)
                self.logger.info(f"Successfully deleted {total_deleted} user(s) for card UID: {card_uid}")
            else:
                self.logger.info(f"No user found to delete for card UID: {card_uid}")

            workbook.close()
            return True

        except Exception as e:
            self.logger.error(f"Error deleting user from Excel: {str(e)}")
            return False

    def _delete_from_worksheet(self, worksheet, card_uid: str) -> int:
        try:
            # Get UID column index
            uid_column_index = self._column_letter_to_index(self.columns['card_uid']) + 1  # openpyxl is 1-based

            # Find and delete matching rows
            rows_to_clear = []

            for row_num in range(1, worksheet.max_row + 1):  # Include all rows
                cell_value = worksheet.cell(row=row_num, column=uid_column_index).value
                if cell_value and str(cell_value).strip().lower() == card_uid.lower():
                    rows_to_clear.append(row_num)

            # Delete rows (from bottom to top to avoid index shifting)
            for row_num in rows_to_clear:
                for col_num in range(1, 12):  # Columns 1 to 11
                    worksheet.cell(row=row_num, column=col_num).value = None
                self.logger.debug(f"Cleared content in first 11 columns of row {row_num} for card UID: {card_uid}")

            return len(rows_to_clear)

        except Exception as e:
            self.logger.error(f"Error deleting from worksheet: {str(e)}")
            return 0

    def get_user_count(self) -> int:
        try:
            total_count = 0
            for worksheet_name in self.config.EXCEL_WORKSHEETS:
                try:
                    df = pd.read_excel(self.excel_path, sheet_name=worksheet_name, header=None)
                    uid_col_idx = self._column_letter_to_index(self.columns['card_uid'])
                    if df.shape[1] > uid_col_idx:
                        non_empty_rows = df.iloc[:, uid_col_idx].notna().sum()
                        total_count += non_empty_rows
                except Exception as e:
                    self.logger.error(f"Error counting users in worksheet '{worksheet_name}': {str(e)}")

            return total_count
        except Exception as e:
            self.logger.error(f"Error counting users in Excel: {str(e)}")
            return 0

    def validate_excel_structure(self) -> bool:

        try:
            workbook = load_workbook(self.excel_path)

            # Check if configured worksheets exist
            missing_worksheets = []
            for worksheet_name in self.config.EXCEL_WORKSHEETS:
                if worksheet_name not in workbook.sheetnames:
                    missing_worksheets.append(worksheet_name)

            if missing_worksheets:
                self.logger.error(f"Excel validation failed. Missing worksheets: {missing_worksheets}")
                self.logger.info(f"Available worksheets: {workbook.sheetnames}")
                return False

            # Check each worksheet for basic structure
            total_rows = 0
            for worksheet_name in self.config.EXCEL_WORKSHEETS:
                try:
                    df = pd.read_excel(self.excel_path, sheet_name=worksheet_name, header=None)

                    # Check if worksheet has enough columns
                    required_cols = [
                        self._column_letter_to_index(self.columns['supervisor']),
                        self._column_letter_to_index(self.columns['gender']),
                        self._column_letter_to_index(self.columns['firstname']),
                        self._column_letter_to_index(self.columns['lastname']),
                        self._column_letter_to_index(self.columns['card_uid'])
                    ]

                    max_required_col = max(required_cols)
                    if df.shape[1] <= max_required_col:
                        self.logger.error(f"Worksheet '{worksheet_name}' doesn't have enough columns. Has {df.shape[1]}, needs {max_required_col + 1}")
                        return False

                    # Count non-empty rows
                    uid_col_idx = self._column_letter_to_index(self.columns['card_uid'])
                    non_empty_rows = df.iloc[:, uid_col_idx].notna().sum()
                    total_rows += non_empty_rows

                    self.logger.info(f"Worksheet '{worksheet_name}' validation successful. Found {non_empty_rows} data rows.")

                except Exception as e:
                    self.logger.error(f"Error validating worksheet '{worksheet_name}': {str(e)}")
                    return False

            workbook.close()

            if total_rows == 0:
                self.logger.warning("Excel file has no data rows across all worksheets")
                return False

            self.logger.info(f"Excel validation successful. Found {total_rows} total users across all worksheets.")
            return True

        except Exception as e:
            self.logger.error(f"Excel validation error: {str(e)}")
            return False

    def test_connection(self) -> bool:
        try:
            return self.validate_excel_structure()
        except Exception as e:
            self.logger.error(f"Excel connection test failed: {str(e)}")
            return False

    def _generate_outlook_address(self, lastname: str, firstname: str) -> str:
        if not lastname and not firstname:
            return ""

        if lastname and firstname:
            return f"{lastname}, {firstname}"
        elif lastname:
            return lastname
        elif firstname:
            return firstname

        return ""

    def _column_letter_to_index(self, letter: str) -> int:
        if len(letter) == 1:
            return ord(letter.upper()) - ord('A')
        else:
            # Handle multi-letter columns like AA, AB, etc.
            result = 0
            for i, char in enumerate(reversed(letter.upper())):
                result += (ord(char) - ord('A') + 1) * (26 ** i)
            return result - 1
